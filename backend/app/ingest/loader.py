"""Multi-format document loaders."""

from __future__ import annotations

import csv
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from app.rag.types import ParsedBlock, ParsedDocument


class DocumentLoader:
    """Load text blocks from PDF, DOCX, TXT, CSV, and XLSX files."""

    def load(self, file_path: Path) -> list[ParsedBlock]:
        """Dispatch to a file-type specific loader."""

        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            return self._load_pdf(file_path)
        if suffix == ".docx":
            return self._load_docx(file_path)
        if suffix == ".txt":
            return self._load_txt(file_path)
        if suffix == ".xlsx":
            return self._load_xlsx(file_path)
        if suffix == ".csv":
            return self._load_csv(file_path)
        raise ValueError(f"暂不支持的文件格式: {suffix}")

    def _load_pdf(self, file_path: Path) -> list[ParsedBlock]:
        """Load PDF pages as blocks."""

        reader = PdfReader(str(file_path))
        blocks: list[ParsedBlock] = []
        for index, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                blocks.append(ParsedBlock(text=text, page=index))
        return blocks


class DocumentParser:
    """Convert source files into normalized parsed documents."""

    def __init__(self, loader: DocumentLoader | None = None) -> None:
        """Initialize the parser."""

        self.loader = loader or DocumentLoader()

    def parse(self, document_id: str, file_path: Path, doc_name: str, doc_type: str) -> ParsedDocument:
        """Parse a source file into normalized blocks with document metadata."""

        blocks = self.loader.load(file_path)
        return ParsedDocument.create(
            doc_id=document_id,
            doc_name=doc_name,
            doc_type=doc_type,
            blocks=blocks,
        )

    def _load_docx(self, file_path: Path) -> list[ParsedBlock]:
        """Load DOCX paragraphs while preserving headings."""

        document = DocxDocument(str(file_path))
        blocks: list[ParsedBlock] = []
        current_section: str | None = None
        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            style_name = paragraph.style.name if paragraph.style is not None else ""
            if style_name.startswith("Heading"):
                current_section = text
                continue
            blocks.append(ParsedBlock(text=text, section=current_section))
        return blocks

    def _load_txt(self, file_path: Path) -> list[ParsedBlock]:
        """Load a plain-text document and split by paragraphs."""

        content = file_path.read_text(encoding="utf-8")
        return [ParsedBlock(text=part.strip()) for part in content.split("\n\n") if part.strip()]

    def _load_xlsx(self, file_path: Path) -> list[ParsedBlock]:
        """Load spreadsheet rows as structured textual blocks."""

        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        blocks: list[ParsedBlock] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            for row in rows[1:]:
                values = []
                for header, cell in zip(headers, row, strict=False):
                    if header and cell not in (None, ""):
                        values.append(f"{header}: {cell}")
                if values:
                    blocks.append(ParsedBlock(text="；".join(values), section=sheet.title))
        return blocks

    def _load_csv(self, file_path: Path) -> list[ParsedBlock]:
        """Load CSV rows as structured textual blocks."""

        blocks: list[ParsedBlock] = []
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                values = [f"{key}: {value}" for key, value in row.items() if value]
                if values:
                    blocks.append(ParsedBlock(text="；".join(values)))
        return blocks
